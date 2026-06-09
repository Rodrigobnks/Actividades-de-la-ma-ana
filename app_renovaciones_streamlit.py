# app_renovaciones_streamlit.py
# App Streamlit para descargar renovaciones, consolidar o ejecutar proceso completo.
# La lógica interna de descarga y consolidado se conserva; la app solo agrega menú y selección de semana.

import streamlit as st
import io
import contextlib
import zipfile
from pathlib import Path
import datetime as _dt

# RUN_RENOMBRAR_MIN.py
FECHA_DESDE, FECHA_HASTA = "08-06-2026", "14-06-2026"
SEMANA_SELECCIONADA = 24
DIA_SELECCIONADO = "Lunes"

RUTA_BASE_DEFAULT = r"C:\Users\EQUIPO\Desktop\Renovaciones"
RUTA_CONSOLIDADO_DEFAULT = r"C:\Users\EQUIPO\Desktop\Renovaciones\Renovacion"

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time, os, glob, datetime, shutil, re, traceback

# ---------- Configuración de Rutas ----------
CHROME_BIN = r"C:\Program Files\Google\Chrome\Application\chrome.exe"
DRIVER_PATH = None

DOWNLOAD_DIR = RUTA_BASE_DEFAULT
CONSOLIDADO_DIR = RUTA_BASE_DEFAULT
ANTERIORES_DIR = RUTA_BASE_DEFAULT

DEFAULT_DOWNLOAD_TIMEOUT = 180
PRESICO_FINAL_TIMEOUT = 600
PRESICO_FINAL_RETRIES = 4

def limpiar_carpeta_descargas():
    os.makedirs(DOWNLOAD_DIR, exist_ok=True)

    for archivo in glob.glob(os.path.join(DOWNLOAD_DIR, "*")):
        try:
            # No borrar la carpeta del consolidado.
            # Las descargas se limpian, pero C:\Users\EQUIPO\Desktop\Renovaciones\Renovacion se conserva.
            if os.path.isdir(archivo) and os.path.basename(archivo).lower() == "renovacion":
                continue
            if os.path.isfile(archivo) or os.path.islink(archivo):
                os.remove(archivo)
            elif os.path.isdir(archivo):
                shutil.rmtree(archivo)
        except Exception as e:
            print(f"⚠️ No se pudo borrar {archivo}: {e}")


PREFS = {
    "download.default_directory": DOWNLOAD_DIR,
    "download.prompt_for_download": False,
    "download.directory_upgrade": True,
    "safebrowsing.enabled": True,
    "profile.default_content_settings.popups": 0
}

SITES = [
    {"name": "EL SALVADOR", "url": "https://front-salvador.caprepaprojects.com/login?returnUrl=%2F", "logos": ["PISTIYO", "moderna salvador"], "week_selection": "mat_select", "week_option_text": "Semana 14"},
    {"name": "COLOMBIA", "url": "https://front-colombia.caprepaprojects.com/login?returnUrl=%2F", "logos": ["PRESICO"], "week_selection": "mat_select", "week_option_text": "Semana 14"},
    {"name": "HONDURAS", "url": "https://front-honduras.caprepaprojects.com/login?returnUrl=%2F", "logos": ["PISTIYO"], "week_selection": "mat_select", "week_option_text": "Semana 14"},
    {"name": "PERU", "url": "https://front-peru.caprepaprojects.com/login?returnUrl=%2F", "logos": ["CASITA", "PRESICO"], "week_selection": "checkbox", "week_checkbox_id": "14-input"},
    {"name": "GUATEMALA", "url": "https://front-guatemala.caprepaprojects.com/login?returnUrl=%2F", "logos": ["CASITA", "PISTIYO", "PRESICO"], "week_selection": "checkbox", "week_checkbox_id": "14-input"},
    {"name": "NICARAGUA", "url": "https://front-nicaragua.caprepaprojects.com/login?returnUrl=%2F", "logos": ["PISTIYO"], "week_selection": "mat_select", "week_option_text": "Semana 14"}
]

# ---------- Estado visual de descargas en Streamlit ----------
ESTADO_DESCARGAS = {}
ESTADO_DESCARGAS_PLACEHOLDER = None


def nombre_marca_visible(logo, site_cfg):
    """Muestra la marca con el mismo nombre final que se usa al renombrar archivos."""
    return map_logo(logo, site_cfg)


def clave_estado(site_cfg, logo):
    return (site_cfg.get("name", ""), nombre_marca_visible(logo, site_cfg))


def actualizar_estado_descarga(site_cfg, logo, estado, detalle=""):
    """Actualiza únicamente el tablero visual; no cambia la lógica de Selenium ni de consolidado."""
    global ESTADO_DESCARGAS, ESTADO_DESCARGAS_PLACEHOLDER

    if not ESTADO_DESCARGAS_PLACEHOLDER:
        return

    pais, marca = clave_estado(site_cfg, logo)
    ESTADO_DESCARGAS[(pais, marca)] = {
        "País": pais,
        "Marca": marca,
        "Estado": estado,
        "Detalle": detalle,
        "Actualizado": datetime.datetime.now().strftime("%H:%M:%S"),
    }

    try:
        df_estado = pd.DataFrame(ESTADO_DESCARGAS.values())
        orden_estado = {
            "🔄 EN PROCESO": 0,
            "⏳ PENDIENTE": 1,
            "✅ DESCARGADO": 2,
            "❌ ERROR": 3,
        }
        df_estado["_orden"] = df_estado["Estado"].map(orden_estado).fillna(9)
        df_estado = df_estado.sort_values(["_orden", "País", "Marca"]).drop(columns=["_orden"])
        ESTADO_DESCARGAS_PLACEHOLDER.dataframe(df_estado, use_container_width=True, hide_index=True)
    except Exception:
        pass


def inicializar_estado_descargas(accion):
    """Prepara el menú/tabla de estado según la acción seleccionada."""
    global ESTADO_DESCARGAS, ESTADO_DESCARGAS_PLACEHOLDER

    ESTADO_DESCARGAS = {}

    if accion not in ("Ejecutar proceso completo", "Solo descargar archivos"):
        ESTADO_DESCARGAS_PLACEHOLDER = None
        return None

    contenedor = st.container()
    with contenedor:
        st.subheader("Estado de descargas por país y marca")
        ESTADO_DESCARGAS_PLACEHOLDER = st.empty()

    # Mismo orden de ejecución que ejecutar_descarga_todo(): primero bloque prioritario y luego PRESICO finales.
    for s in SITES:
        revisar_pausa()
        marcas_filtradas = [
            m for m in s["logos"]
            if not (
                (s["name"] == "GUATEMALA" or s["name"] == "PERU")
                and m == "PRESICO"
            )
        ]
        for logo in marcas_filtradas:
            actualizar_estado_descarga(s, logo, "⏳ PENDIENTE", "En espera")

    for pais in ["GUATEMALA", "PERU"]:
        revisar_pausa()
        cfg = next((s for s in SITES if s["name"] == pais), None)
        if cfg:
            actualizar_estado_descarga(cfg, "PRESICO", "⏳ PENDIENTE", "En espera")

    return contenedor


# ---------- Control de pausa / reanudar ----------
def proceso_pausado():
    try:
        return bool(st.session_state.get("pausar_proceso_renovaciones", False))
    except Exception:
        return False


def revisar_pausa():
    """Pausa el proceso entre pasos sin cambiar la lógica interna de descarga."""
    try:
        aviso = st.empty()
        while st.session_state.get("pausar_proceso_renovaciones", False):
            aviso.warning("⏸️ Proceso pausado. Presiona Reanudar proceso para continuar.")
            time.sleep(1)
        aviso.empty()
    except Exception:
        pass


def dormir_controlado(segundos):
    """Respeta los mismos tiempos, pero permite pausar entre intervalos cortos."""
    fin = time.time() + segundos
    while time.time() < fin:
        revisar_pausa()
        time.sleep(min(0.5, max(0, fin - time.time())))


USERNAME, PASSWORD = "rodrigoseas", "Saldaa6103"

LOGO_XPATHS = {
    ("GUATEMALA", "PRESICO"): "//img[contains(@class,'btn-logo') and contains(@src,'logo_guatemala.png') and @title='GUATEMALA']",
    ("GUATEMALA", "CASITA"): "//img[contains(@class,'btn-logo') and contains(@src,'logo_casita.png') and @title='PRESTAMOS LA CASITA']",
    ("GUATEMALA", "CAPREPA"): "//img[contains(@class,'btn-logo') and contains(@src,'logo_caprepa.png') and @title='Corporativo Caprepa']",
    ("GUATEMALA", "PISTIYO"): "//img[contains(@class,'btn-logo') and contains(translate(@src,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'pistiyo')]",

    ("PERU", "PRESICO"): "//img[contains(@class,'btn-logo') and (contains(translate(@src,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'presico') or contains(translate(@src,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'peru'))]",
    ("PERU", "CASITA"): "//img[contains(@class,'btn-logo') and contains(translate(@src,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'casita')]",
}

# ---------- Helpers ----------
def create_driver():
    opts = Options()
    opts.binary_location = CHROME_BIN
    opts.add_argument("--start-maximized")
    opts.add_experimental_option("prefs", PREFS)
    opts.add_argument("--log-level=3")

    try:
        if DRIVER_PATH and os.path.exists(DRIVER_PATH):
            service = Service(executable_path=DRIVER_PATH)
        else:
            try:
                path = ChromeDriverManager().install()
                service = Service(path)
            except Exception:
                print("⚠️ No se pudo usar DriverManager. Usando driver del sistema...")
                return webdriver.Chrome(options=opts)

        return webdriver.Chrome(service=service, options=opts)

    except Exception as e:
        print(f"❌ Error crítico al crear driver: {e}")
        return None


def wait(driver, cond, timeout=20):
    return WebDriverWait(driver, timeout).until(cond)


def sanitize(name):
    return re.sub(r'[<>:"/\\|?*]+', "", " ".join(name.split()))


def map_logo(logo, site):
    L = logo.strip().lower()
    pais = site.get("name", "").upper()

    if pais == "EL SALVADOR" and "moderna" in L:
        return "LA MODERNA"
    if pais in ("PERU", "GUATEMALA") and "casita" in L:
        return "LA CASITA"
    if pais == "GUATEMALA" and (L == "guatemala" or L == "presico"):
        return "PRESICO"

    return logo.strip().upper()


def list_files(pattern="*"):
    return [
        p for p in glob.glob(os.path.join(DOWNLOAD_DIR, pattern))
        if os.path.isfile(p)
    ]


def newest_file(exclude: set, since_ts: float):
    cand = [p for p in list_files("*") if p not in exclude]
    cand = [p for p in cand if os.path.getmtime(p) + 0.05 >= since_ts]

    cand = [
        p for p in cand
        if not p.endswith(".crdownload")
        and not p.endswith(".tmp")
        and os.path.getsize(p) > 0
    ]

    if not cand:
        return None

    return max(cand, key=os.path.getmtime)


def safe_move_or_rename(src, dst):
    base, ext = os.path.splitext(dst)
    i = 1
    target = dst

    while os.path.exists(target):
        target = f"{base}_{i}{ext}"
        i += 1

    shutil.move(src, target)
    return target


def ensure(d):
    os.makedirs(d, exist_ok=True)


# ---------- Renombrado y movimiento ----------
def rename_last_and_move(logo_title, site_cfg, before_files, ts_inicio):
    f = newest_file(before_files, ts_inicio)

    if not f:
        print("⚠️ No se detectó archivo nuevo.")
        return None

    try:
        mdate = datetime.date.fromtimestamp(os.path.getmtime(f)) - datetime.timedelta(days=1)
    except:
        try:
            dd, mm, yy = FECHA_HASTA.split("-")
            mdate = datetime.date(int(yy), int(mm), int(dd)) - datetime.timedelta(days=1)
        except:
            mdate = None

    fecha_str = mdate.strftime("%d_%m_%Y") if mdate else FECHA_HASTA.replace("-", "_")

    nuevo = f"Renovaciones_{fecha_str} {map_logo(logo_title, site_cfg)} {site_cfg.get('name', '')}"
    nuevo = sanitize(nuevo) + os.path.splitext(f)[1]

    dst = os.path.join(DOWNLOAD_DIR, nuevo)

    try:
        renamed = safe_move_or_rename(f, dst)
        print("✅ Renombrado a:", os.path.basename(renamed))
        move_to_consolidado(renamed)
        return renamed

    except Exception as e:
        print("⚠️ Error renombrando:", e)
        return None


def move_to_consolidado(file_path):
    try:
        ensure(CONSOLIDADO_DIR)

        fname = os.path.basename(file_path)
        dst = os.path.join(CONSOLIDADO_DIR, fname)

        if os.path.abspath(file_path) == os.path.abspath(dst):
            print("✅ Archivo final guardado en:", dst)
            return

        safe_move_or_rename(file_path, dst)
        print("✅ Archivo final guardado en:", dst)

    except Exception as e:
        print("⚠️ Error moviendo archivo final:", e)


# ---------- UI Actions ----------
def try_find(driver, title, site=None):
    pais = site.get("name", "").upper() if site else ""
    marca = title.strip().upper()

    xpath_mapeado = LOGO_XPATHS.get((pais, marca))

    if xpath_mapeado:
        try:
            el = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, xpath_mapeado))
            )
            return el
        except:
            print(f"⚠️ No se encontró con mapa exacto: {pais} - {marca}")

    checks = [
        (By.XPATH, f"//img[@title='{title}']"),
        (By.XPATH, f"//img[@alt='{title}']"),
        (By.XPATH, f"//button[normalize-space()='{title}']"),
        (By.XPATH, f"//a[normalize-space()='{title}']"),
    ]

    for by, sel in checks:
        try:
            el = WebDriverWait(driver, 6).until(
                EC.element_to_be_clickable((by, sel))
            )
            return el
        except:
            continue

    return None


def seleccionar_semana(driver, cfg):
    try:
        if cfg["week_selection"] == "checkbox":
            cb = wait(
                driver,
                EC.element_to_be_clickable((By.ID, cfg["week_checkbox_id"])),
                12
            )

            if not cb.is_selected():
                driver.execute_script("arguments[0].click();", cb)

        else:
            texto_semana = cfg.get("week_option_text", "Semana 14")

            try:
                mat_select = wait(
                    driver,
                    EC.element_to_be_clickable((By.XPATH, "//mat-select[contains(@class,'mat-select')]")),
                    12
                )
                driver.execute_script("arguments[0].scrollIntoView({block:'center'});", mat_select)
                time.sleep(0.5)
                try:
                    mat_select.click()
                except:
                    driver.execute_script("arguments[0].click();", mat_select)
            except:
                driver.execute_script("document.querySelectorAll('mat-select')[0].click();")

            time.sleep(1)

            # Igual que el flujo original: abrir el mat-select y elegir Semana 14.
            # Se hace clic sobre el mat-option completo, no solo sobre el span,
            # porque en Angular a veces el span abre visualmente pero no registra la selección.
            opt = wait(
                driver,
                EC.element_to_be_clickable(
                    (By.XPATH, f"//mat-option[.//span[contains(@class,'mat-option-text') and normalize-space()='{texto_semana}']]")
                ),
                12
            )

            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", opt)
            time.sleep(0.5)

            try:
                opt.click()
            except:
                driver.execute_script("arguments[0].click();", opt)

            time.sleep(1)

            # Cierre/confirmación visual del desplegable antes de generar reporte.
            try:
                wait(
                    driver,
                    EC.invisibility_of_element_located((By.XPATH, "//div[contains(@class,'cdk-overlay-pane')]//mat-option")),
                    5
                )
            except:
                pass

        print("✅ Semana seleccionada")

    except Exception as e:
        print("⚠️ Semana:", e)

def esperar_descarga(before, ts, timeout):
    start = time.time()

    while time.time() - start < timeout:
        revisar_pausa()
        time.sleep(1)
        cand = newest_file(before, ts)

        if cand and os.path.isfile(cand):
            return cand

    return None


def es_presico_final(site_cfg, logo):
    pais = site_cfg.get("name", "").upper()
    marca = logo.strip().upper()
    return pais in ("GUATEMALA", "PERU") and marca == "PRESICO"


def generar_reporte(driver, logo, site_cfg):
    try:
        intentos = PRESICO_FINAL_RETRIES if es_presico_final(site_cfg, logo) else 1
        timeout_descarga = PRESICO_FINAL_TIMEOUT if es_presico_final(site_cfg, logo) else DEFAULT_DOWNLOAD_TIMEOUT

        for intento in range(1, intentos + 1):
            revisar_pausa()
            actualizar_estado_descarga(site_cfg, logo, "🔄 EN PROCESO", f"Intento {intento}/{intentos}")
            print(f"\n🔁 Intento {intento}/{intentos} para {logo} en {site_cfg['name']}")

            el = try_find(driver, logo, site_cfg)

            if not el:
                print(f"❌ No se encontró logo {logo} en {site_cfg['name']}")
                actualizar_estado_descarga(site_cfg, logo, "❌ ERROR", "No se encontró logo")
                return

            driver.execute_script("arguments[0].scrollIntoView(true);", el)

            try:
                el.click()
            except:
                driver.execute_script("arguments[0].click();", el)

            time.sleep(2)

            try:
                btn = WebDriverWait(driver, 3).until(
                    EC.element_to_be_clickable(
                        (By.XPATH, "//button[normalize-space(text())='Aceptar']")
                    )
                )
                driver.execute_script("arguments[0].click();", btn)
                time.sleep(1)
            except:
                pass

            try:
                menu = wait(
                    driver,
                    EC.element_to_be_clickable(
                        (By.XPATH, "//fa-icon[@class='ng-fa-icon']/*[@data-icon='bars']")
                    ),
                    8
                )

                try:
                    menu.click()
                except:
                    driver.execute_script("arguments[0].click();", menu)

            except:
                try:
                    mf = wait(
                        driver,
                        EC.element_to_be_clickable(
                            (By.XPATH, "//span[normalize-space(text())='Menu' or normalize-space(text())='MENU']")
                        ),
                        6
                    )
                    driver.execute_script("arguments[0].click();", mf)
                except:
                    pass

            time.sleep(1)

            try:
                rep = wait(
                    driver,
                    EC.element_to_be_clickable(
                        (By.XPATH, "//span[normalize-space(text())='Reportes']")
                    ),
                    8
                )
                rep.click()
            except:
                try:
                    driver.execute_script(
                        "arguments[0].click();",
                        wait(
                            driver,
                            EC.element_to_be_clickable(
                                (By.XPATH, "//*[contains(translate(text(),'REPORTES','reportes'),'reportes')]")
                            ),
                            6
                        )
                    )
                except:
                    pass

            time.sleep(1)

            renov = wait(
                driver,
                EC.presence_of_element_located(
                    (By.XPATH, "//span[normalize-space(text())='Renovaciones']")
                ),
                12
            )

            driver.execute_script("arguments[0].click();", renov)
            time.sleep(1)

            fd = wait(
                driver,
                EC.element_to_be_clickable(
                    (By.XPATH, "//input[@placeholder='Fecha Desde' and @type='date']")
                ),
                12
            )
            fd.clear()
            fd.send_keys(FECHA_DESDE)

            fh = wait(
                driver,
                EC.element_to_be_clickable(
                    (By.XPATH, "//input[@placeholder='Fecha Hasta' and @type='date']")
                ),
                12
            )
            fh.clear()
            fh.send_keys(FECHA_HASTA)

            time.sleep(1)

            seleccionar_semana(driver, site_cfg)
            time.sleep(1)

            try:
                r = wait(
                    driver,
                    EC.presence_of_element_located(
                        (By.XPATH, "//div[contains(@class,'mat-radio-outer-circle')]")
                    ),
                    5
                )
                driver.execute_script("arguments[0].click();", r)
            except:
                pass

            before = set(list_files("*"))
            ts = time.time()

            gen = wait(
                driver,
                EC.element_to_be_clickable(
                    (By.XPATH, "//button[@title='Generar Reporte']")
                ),
                12
            )

            driver.execute_script("arguments[0].click();", gen)

            print(f"⏳ Generando para {logo} en {site_cfg['name']}... Tiempo máximo: {timeout_descarga} segundos")

            newf = esperar_descarga(before, ts, timeout_descarga)

            if not newf:
                print(f"⚠️ No se generó archivo para {logo} en {site_cfg['name']} en el intento {intento}.")

                if intento < intentos:
                    try:
                        home = wait(
                            driver,
                            EC.element_to_be_clickable(
                                (By.XPATH, "//a[contains(@class,'nav-item home')]//label[normalize-space()='INICIO']")
                            ),
                            6
                        )
                        driver.execute_script("arguments[0].click();", home)
                        time.sleep(2)
                    except:
                        driver.get(site_cfg["url"])
                        time.sleep(3)

                    continue

                print("❌ Se agotaron los intentos.")
                actualizar_estado_descarga(site_cfg, logo, "❌ ERROR", "No se generó archivo")
                return

            print("✅ Descarga encontrada:", os.path.basename(newf))

            renamed = rename_last_and_move(logo, site_cfg, before, ts)

            if renamed:
                actualizar_estado_descarga(site_cfg, logo, "✅ DESCARGADO", os.path.basename(renamed))
                return

        time.sleep(1)

    except Exception as e:
        actualizar_estado_descarga(site_cfg, logo, "❌ ERROR", str(e)[:120])
        print("❌ generar_reporte error:", e)
        traceback.print_exc()


# ---------- Lógica Principal ----------
def ejecutar_sitio(site_cfg, marcas_especificas=None):
    print("\n" + "=" * 40)
    print(f" PROCESANDO: {site_cfg['name']}")
    print("=" * 40)

    driver = None

    try:
        driver = create_driver()

        if not driver:
            marcas = marcas_especificas if marcas_especificas else site_cfg["logos"]
            for logo in marcas:
                actualizar_estado_descarga(site_cfg, logo, "❌ ERROR", "Falló driver")
            return "❌ FALLÓ DRIVER"

        driver.get(site_cfg["url"])
        wait(driver, EC.presence_of_element_located((By.TAG_NAME, "body")), 20)
        time.sleep(2)

        try:
            u = wait(
                driver,
                EC.presence_of_element_located(
                    (By.XPATH, "//input[@formcontrolname='username']")
                ),
                8
            )
            u.clear()
            u.send_keys(USERNAME)

            p = driver.find_element(By.XPATH, "//input[@formcontrolname='password']")
            p.clear()
            p.send_keys(PASSWORD)
            p.send_keys(Keys.RETURN)

            wait(driver, EC.presence_of_element_located((By.TAG_NAME, "body")), 10)
            time.sleep(2)

        except:
            print("⚠️ Login omitido o sesión ya activa.")

        marcas = marcas_especificas if marcas_especificas else site_cfg["logos"]

        for logo in marcas:
            revisar_pausa()
            generar_reporte(driver, logo, site_cfg)

            try:
                home = wait(
                    driver,
                    EC.element_to_be_clickable(
                        (By.XPATH, "//a[contains(@class,'nav-item home')]//label[normalize-space()='INICIO']")
                    ),
                    6
                )
                driver.execute_script("arguments[0].click();", home)
                time.sleep(1)

            except:
                pass

        return "✅ ÉXITO"

    except Exception as e:
        marcas = marcas_especificas if marcas_especificas else site_cfg["logos"]
        for logo in marcas:
            pais, marca = clave_estado(site_cfg, logo)
            if ESTADO_DESCARGAS.get((pais, marca), {}).get("Estado") != "✅ DESCARGADO":
                actualizar_estado_descarga(site_cfg, logo, "❌ ERROR", str(e)[:120])
        print(f"❌ Error en {site_cfg['name']}: {e}")
        return "❌ FALLÓ"

    finally:
        if driver:
            try:
                driver.quit()
            except:
                pass

        time.sleep(1)

# ============================================================
# BLOQUE CONSOLIDADO ORIGINAL
# ============================================================
import pandas as pd
from pathlib import Path
import re
import unicodedata

from openpyxl import load_workbook
from openpyxl.styles import Font


# ============================================================
# CONFIGURACIÓN
# ============================================================

CARPETA_ENTRADA = Path(RUTA_BASE_DEFAULT)
CARPETA_SALIDA = Path(RUTA_CONSOLIDADO_DEFAULT)

ARCHIVO_SALIDA = CARPETA_SALIDA / f"Semana {SEMANA_SELECCIONADA}.xlsx"


# ============================================================
# FUNCIONES AUXILIARES
# ============================================================

def limpiar_texto(valor):
    if pd.isna(valor):
        return valor

    texto = str(valor)
    texto = "".join(ch for ch in texto if unicodedata.category(ch)[0] != "C")
    return texto.strip()


def normalizar_columna(columna):
    """
    Normaliza nombres de columnas para evitar errores por:
    - Mayúsculas/minúsculas
    - Espacios
    - Guiones
    - Acentos
    - Caracteres invisibles
    """
    texto = str(columna).strip().lower()

    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(c for c in texto if not unicodedata.combining(c))

    texto = texto.replace(" ", "_")
    texto = texto.replace("-", "_")
    texto = texto.replace(".", "_")
    texto = texto.replace("/", "_")

    texto = re.sub(r"[^a-z0-9_]", "", texto)
    texto = re.sub(r"_+", "_", texto)

    return texto.strip("_")


def estandarizar_columnas(df):
    """
    Renombra columnas aunque vengan con nombres ligeramente distintos.
    Esto corrige especialmente fecha_nuevo_desembolso.
    """

    df.columns = [normalizar_columna(c) for c in df.columns]

    equivalencias = {
        "zonas": [
            "zonas",
            "zona",
        ],
        "ruta": [
            "ruta",
        ],
        "coordinadora_id": [
            "coordinadora_id",
            "id_coordinadora",
            "coordinadora",
            "coord_id",
        ],
        "localidad": [
            "localidad",
        ],
        "cliente_id": [
            "cliente_id",
            "id_cliente",
            "dbc_cliente_id",
        ],
        "cliente": [
            "cliente",
            "nombre_cliente",
        ],
        "celular_cliente": [
            "celular_cliente",
            "celular",
            "telefono_celular_cliente",
        ],
        "telefono_casa_cliente": [
            "telefono_casa_cliente",
            "telefono_casa",
            "tel_casa_cliente",
        ],
        "estatus_des": [
            "estatus_des",
        ],
        "estatus_desembolso": [
            "estatus_desembolso",
            "estatus_de_desembolso",
            "estado_desembolso",
        ],
        "monto_desembolso": [
            "monto_desembolso",
            "monto",
            "importe_desembolso",
        ],
        "numero_pago": [
            "numero_pago",
            "num_pago",
            "no_pago",
        ],
        "fecha_pago": [
            "fecha_pago",
            "fecha_de_pago",
        ],
        "fecha_pago_final": [
            "fecha_pago_final",
            "fecha_final_pago",
            "fecha_de_pago_final",
        ],
        "fecha_nuevo_desembolso": [
            "fecha_nuevo_desembolso",
            "fecha_nuevo_desembolso_",
            "fecha_de_nuevo_desembolso",
            "fecha_nuevo_desemb",
            "fecha_desembolso_nuevo",
            "fecha_nvo_desembolso",
            "fecha_nuevo_desembolso_final",
            "fecha_nuevo_desembolso_real",
            "fecha_nueva_desembolso",
            "fecha_nuevo_desem",
            "fecha_nuevo_des",
            "fecha_nuevo",
            "fecha_nuevo_desembolso_1",
            "fecha_nuevo_desembolso2",
            "fecha_nuevo_desembolso_2",
        ],
        "semana_mes": [
            "semana_mes",
            "semana_del_mes",
        ],
        "vuelta": [
            "vuelta",
        ],
        "mora_total": [
            "mora_total",
            "mora",
        ],
        "entregado": [
            "entregado",
        ],
        "tipo_desembolso": [
            "tipo_desembolso",
            "tipo_de_desembolso",
        ],
        "tipo_desembolso_nuevo": [
            "tipo_desembolso_nuevo",
            "tipo_de_desembolso_nuevo",
            "tipo_desembolso_nvo",
        ],
    }

    renombres = {}
    columnas_actuales = set(df.columns)

    for nombre_final, posibles in equivalencias.items():
        for posible in posibles:
            posible_norm = normalizar_columna(posible)

            if posible_norm in columnas_actuales:
                renombres[posible_norm] = nombre_final
                break

    df = df.rename(columns=renombres)

    return df


def leer_csv_seguro(ruta_archivo):
    """
    Lee CSV con varias codificaciones.

    Nota:
    Se usa engine="python" porque permite detectar separador con sep=None.
    Por eso NO se usa low_memory=False, porque pandas no lo permite con engine="python".
    """

    codificaciones = ["utf-8-sig", "utf-8", "latin1", "cp1252"]
    ultimo_error = None

    for encoding in codificaciones:
        try:
            df = pd.read_csv(
                ruta_archivo,
                encoding=encoding,
                dtype=str,
                sep=None,
                engine="python"
            )

            df = estandarizar_columnas(df)
            return df

        except UnicodeDecodeError as e:
            ultimo_error = e
            continue

        except Exception as e:
            ultimo_error = e
            continue

    raise ValueError(
        f"No se pudo leer el archivo: {ruta_archivo.name}. "
        f"Último error: {ultimo_error}"
    )


def extraer_fecha_corte(nombre_archivo):
    """
    Extrae fecha desde nombres tipo:
    Renovaciones_07_06_2026 LA CASITA GUATEMALA.csv
    """

    match = re.search(r"Renovaciones_(\d{2})_(\d{2})_(\d{4})", nombre_archivo)

    if not match:
        return pd.NaT

    dia, mes, anio = match.groups()

    return pd.to_datetime(
        f"{dia}/{mes}/{anio}",
        dayfirst=True,
        errors="coerce"
    )


def quitar_prefijo_fecha(nombre_archivo):
    return re.sub(
        r"^Renovaciones_\d{2}_\d{2}_\d{4}\s*",
        "",
        nombre_archivo
    )


def extraer_unidad_negocio(nombre_archivo):
    """
    Extrae unidad de negocio desde el nombre del archivo.
    Replica la lógica del Power Query.
    """

    texto = quitar_prefijo_fecha(nombre_archivo)

    reemplazos = [
        " GUATEMALA.csv",
        "PERU.csv",
        " EL SALVADOR.csv",
        " HONDURAS.csv",
        "NICARAGUA.csv",
        " COLOMBIA.csv",
    ]

    for r in reemplazos:
        texto = texto.replace(r, "")

    texto = texto.replace("PISTIYO EL SALVADOR.csv", "PISTIYO")
    texto = texto.replace(".csv", "")

    return limpiar_texto(texto)


def extraer_pais(nombre_archivo):
    """
    Extrae país desde el nombre del archivo.
    Replica la lógica del Power Query.
    """

    texto = quitar_prefijo_fecha(nombre_archivo)

    reemplazos = [
        "LA CASITA ",
        "PRESICO",
        "LA MODERNA ",
        "PISTIYO ",
        ".csv",
    ]

    for r in reemplazos:
        texto = texto.replace(r, "")

    texto = texto.replace(" ", "")

    reemplazos_pais = {
        "PISTIYOELSALVADOR": "EL SALVADOR",
        "ELSALVADOR": "EL SALVADOR",
    }

    texto = reemplazos_pais.get(texto, texto)

    return limpiar_texto(texto)


def convertir_numero(serie):
    return pd.to_numeric(serie, errors="coerce")


def convertir_fecha(serie):
    """
    Convierte fechas sin perder datos.
    Intenta primero formato día/mes/año y luego año/mes/día.
    """

    serie = serie.replace(
        ["", " ", "nan", "NaN", "None", "NULL", "null"],
        pd.NA
    )

    fecha_1 = pd.to_datetime(serie, errors="coerce", dayfirst=True)
    fecha_2 = pd.to_datetime(serie, errors="coerce", dayfirst=False)

    return fecha_1.fillna(fecha_2)


# ============================================================
# FORMATO EXCEL
# ============================================================

def aplicar_formato_excel(ruta_excel):
    wb = load_workbook(ruta_excel)
    ws = wb["Renovacion"]

    fuente_header = Font(name="Century Gothic", size=8, bold=True)
    fuente_body = Font(name="Century Gothic", size=8, bold=False)

    for row in ws.iter_rows():
        for cell in row:
            if cell.row == 1:
                cell.font = fuente_header
            else:
                cell.font = fuente_body

    # Formato de fechas:
    # L = fecha_pago
    # M = fecha_pago_final
    # N = fecha_nuevo_desembolso
    # X = FECHA CORTE
    columnas_fecha = ["L", "M", "N", "X"]

    for col in columnas_fecha:
        for row in range(2, ws.max_row + 1):
            ws[f"{col}{row}"].number_format = "mm-dd-yy"

    anchos = {
        "A": 14.66,   # Pais
        "B": 13.00,   # unidad de negocio
        "C": 23.89,   # zonas
        "D": 22.44,   # ruta
        "E": 13.00,   # coordinadora_id
        "F": 42.89,   # localidad
        "G": 7.89,    # cliente_id
        "H": 42.33,   # cliente
        "I": 15.33,   # estatus_desembolso
        "J": 14.66,   # monto_desembolso
        "K": 10.89,   # numero_pago
        "L": 11.89,   # fecha_pago
        "M": 15.89,   # fecha_pago_final
        "N": 19.55,   # fecha_nuevo_desembolso
        "O": 10.33,   # semana_mes
        "P": 5.33,    # vuelta
        "Q": 10.66,   # mora_total
        "R": 8.33,    # entregado
        "S": 15.33,   # tipo_desembolso
        "T": 20.44,   # tipo_desembolso_nuevo
        "U": 15.66,   # Etiqueta
        "V": 14.89,   # Vencimiento
        "W": 13.00,   # Renovado
        "X": 10.66,   # FECHA CORTE
    }

    for col, width in anchos.items():
        ws.column_dimensions[col].width = width

    wb.save(ruta_excel)


# ============================================================
# PROCESO PRINCIPAL
# ============================================================

def es_lunes_consolidado():
    return str(DIA_SELECCIONADO).strip().lower() == "lunes"


def guardar_o_anexar_consolidado(df):
    """
    Lunes: genera el archivo semanal desde cero.
    Martes a domingo: abre el archivo semanal existente y pega abajo la nueva información.
    No cambia la estructura, fórmulas ni formato del consolidado.
    """
    CARPETA_SALIDA.mkdir(parents=True, exist_ok=True)

    if es_lunes_consolidado():
        print(f"Día seleccionado: {DIA_SELECCIONADO}. Se generará el archivo consolidado semanal desde cero.")
        print("Guardando Excel...")

        with pd.ExcelWriter(
            ARCHIVO_SALIDA,
            engine="openpyxl",
            date_format="mm-dd-yy",
            datetime_format="mm-dd-yy"
        ) as writer:
            df.to_excel(writer, index=False, sheet_name="Renovacion")

        wb = load_workbook(ARCHIVO_SALIDA)
        ws = wb["Renovacion"]

        ultima_fila = ws.max_row
        for fila in range(2, ultima_fila + 1):
            ws[f"U{fila}"] = f'=IF(N{fila}="","Desatendido","Renovado")'
            ws[f"V{fila}"] = 1
            ws[f"W{fila}"] = f'=IF(U{fila}="Renovado",1,0)'

        wb.save(ARCHIVO_SALIDA)
        aplicar_formato_excel(ARCHIVO_SALIDA)
        return

    print(f"Día seleccionado: {DIA_SELECCIONADO}. Se abrirá el archivo semanal existente y se pegará abajo la nueva información.")

    if not ARCHIVO_SALIDA.exists():
        raise FileNotFoundError(
            f"No existe el consolidado semanal para anexar: {ARCHIVO_SALIDA}. "
            "Para crear el archivo desde cero selecciona Lunes."
        )

    wb = load_workbook(ARCHIVO_SALIDA)
    if "Renovacion" not in wb.sheetnames:
        raise ValueError(f"El archivo {ARCHIVO_SALIDA} no contiene la hoja 'Renovacion'.")

    ws = wb["Renovacion"]
    fila_inicio = ws.max_row + 1
    print(f"Pegando {len(df):,} filas nuevas desde la fila {fila_inicio}.")

    for valores in df.itertuples(index=False, name=None):
        ws.append(list(valores))

    ultima_fila = ws.max_row
    for fila in range(fila_inicio, ultima_fila + 1):
        ws[f"U{fila}"] = f'=IF(N{fila}="","Desatendido","Renovado")'
        ws[f"V{fila}"] = 1
        ws[f"W{fila}"] = f'=IF(U{fila}="Renovado",1,0)'

    wb.save(ARCHIVO_SALIDA)
    aplicar_formato_excel(ARCHIVO_SALIDA)


def ejecutar_consolidado():
    print("Leyendo archivos desde:")
    print(CARPETA_ENTRADA)

    if not CARPETA_ENTRADA.exists():
        raise FileNotFoundError(f"No existe la carpeta: {CARPETA_ENTRADA}")

    CARPETA_SALIDA.mkdir(parents=True, exist_ok=True)

    archivos = [
        archivo for archivo in CARPETA_ENTRADA.iterdir()
        if archivo.is_file()
        and archivo.suffix.lower() == ".csv"
        and not archivo.name.startswith(".")
        and not archivo.name.startswith("~$")
    ]

    if not archivos:
        raise FileNotFoundError("No se encontraron archivos CSV en la carpeta.")

    bases = []

    for archivo in archivos:
        print(f"Leyendo: {archivo.name}")

        df_archivo = leer_csv_seguro(archivo)
        df_archivo["Source.Name"] = archivo.name

        if "fecha_nuevo_desembolso" in df_archivo.columns:
            no_vacios = (
                df_archivo["fecha_nuevo_desembolso"]
                .replace("", pd.NA)
                .notna()
                .sum()
            )
            print(f"  fecha_nuevo_desembolso detectada: {no_vacios:,} datos no vacíos")
        else:
            print("  ADVERTENCIA: No se detectó fecha_nuevo_desembolso en este archivo.")

        bases.append(df_archivo)

    df = pd.concat(bases, ignore_index=True)

    print(f"Filas iniciales combinadas: {len(df):,}")

    # ========================================================
    # ASEGURAR COLUMNAS ESPERADAS
    # ========================================================

    columnas_base = [
        "zonas",
        "ruta",
        "coordinadora_id",
        "localidad",
        "cliente_id",
        "cliente",
        "celular_cliente",
        "telefono_casa_cliente",
        "estatus_des",
        "estatus_desembolso",
        "monto_desembolso",
        "numero_pago",
        "fecha_pago",
        "fecha_pago_final",
        "fecha_nuevo_desembolso",
        "semana_mes",
        "vuelta",
        "mora_total",
        "entregado",
        "tipo_desembolso",
        "tipo_desembolso_nuevo",
        "Source.Name",
    ]

    for col in columnas_base:
        if col not in df.columns:
            df[col] = pd.NA

    print("Total no vacío en fecha_nuevo_desembolso antes de convertir:")
    print(df["fecha_nuevo_desembolso"].replace("", pd.NA).notna().sum())

    # ========================================================
    # TIPOS DE DATOS
    # ========================================================

    columnas_numericas = [
        "coordinadora_id",
        "cliente_id",
        "monto_desembolso",
        "numero_pago",
        "semana_mes",
        "vuelta",
        "mora_total",
        "entregado",
    ]

    for col in columnas_numericas:
        df[col] = convertir_numero(df[col])

    columnas_fecha = [
        "fecha_pago",
        "fecha_pago_final",
        "fecha_nuevo_desembolso",
    ]

    for col in columnas_fecha:
        df[col] = convertir_fecha(df[col])

    print("Total no vacío en fecha_nuevo_desembolso después de convertir:")
    print(df["fecha_nuevo_desembolso"].notna().sum())

    # ========================================================
    # CAMPOS DESDE EL NOMBRE DEL ARCHIVO
    # ========================================================

    df["FECHA CORTE"] = df["Source.Name"].apply(extraer_fecha_corte)
    df["unidad de negocio"] = df["Source.Name"].apply(extraer_unidad_negocio)
    df["Pais"] = df["Source.Name"].apply(extraer_pais)

    df["Pais"] = df["Pais"].astype(str).str.replace(" ", "", regex=False)

    df["Pais"] = df["Pais"].replace({
        "PISTIYOELSALVADOR": "EL SALVADOR",
        "ELSALVADOR": "EL SALVADOR",
    })

    df["unidad de negocio"] = df["unidad de negocio"].apply(limpiar_texto)

    # ========================================================
    # LIMPIEZA Y FILTROS COMO POWER QUERY
    # ========================================================

    df["estatus_desembolso"] = df["estatus_desembolso"].astype(str).str.strip()
    df["tipo_desembolso_nuevo"] = df["tipo_desembolso_nuevo"].astype(str).str.strip()

    df = df[
        (df["estatus_desembolso"] == "FINALIZADO")
        & (df["tipo_desembolso_nuevo"] != "Reestructura")
    ].copy()

    print(f"Filas después de filtros: {len(df):,}")

    print("Total no vacío en fecha_nuevo_desembolso después de filtros:")
    print(df["fecha_nuevo_desembolso"].notna().sum())

    # ========================================================
    # COLUMNAS CALCULADAS DEL EXCEL EJEMPLO
    # ========================================================

    df["Etiqueta"] = ""
    df["Vencimiento"] = 1
    df["Renovado"] = ""

    # ========================================================
    # ORDEN FINAL EXACTO
    # ========================================================

    columnas_finales = [
        "Pais",
        "unidad de negocio",
        "zonas",
        "ruta",
        "coordinadora_id",
        "localidad",
        "cliente_id",
        "cliente",
        "estatus_desembolso",
        "monto_desembolso",
        "numero_pago",
        "fecha_pago",
        "fecha_pago_final",
        "fecha_nuevo_desembolso",
        "semana_mes",
        "vuelta",
        "mora_total",
        "entregado",
        "tipo_desembolso",
        "tipo_desembolso_nuevo",
        "Etiqueta",
        "Vencimiento",
        "Renovado",
        "FECHA CORTE",
    ]

    df = df[columnas_finales]

    # ========================================================
    # LIMPIAR NULOS ANTES DE EXPORTAR
    # ========================================================

    # Mantiene la lógica igual, pero evita que pandas exporte el texto "nan"
    # en tipo_desembolso_nuevo. Donde iba "nan", Excel queda en blanco.
    df["tipo_desembolso_nuevo"] = df["tipo_desembolso_nuevo"].replace(
        ["nan", "NaN", "None", "NULL", "null", "<NA>"],
        ""
    )

    df = df.where(pd.notna(df), None)

    # ========================================================
    # GUARDAR O ANEXAR EXCEL SEGÚN EL DÍA SELECCIONADO
    # ========================================================

    guardar_o_anexar_consolidado(df)

    print("Proceso terminado correctamente.")
    print(f"Archivo actualizado en: {ARCHIVO_SALIDA}")

# ============================================================
# ENVOLTURA STREAMLIT: MENÚ + SEMANA
# ============================================================

SEMANA_BASE = 24
FECHA_BASE_DESDE = _dt.date(2026, 6, 8)   # Semana 24: 08-06-2026


def calcular_rango_semana(numero_semana: int):
    """Semana 24 = 08-06-2026 a 14-06-2026; las demás avanzan o retroceden de 7 en 7."""
    inicio = FECHA_BASE_DESDE + _dt.timedelta(days=(int(numero_semana) - SEMANA_BASE) * 7)
    fin = inicio + _dt.timedelta(days=6)
    return inicio, fin


def aplicar_semana_a_config(numero_semana: int):
    """
    Actualiza únicamente FECHA_DESDE y FECHA_HASTA.

    IMPORTANTE:
    La selección de semana dentro del portal NO se modifica aquí.
    Se mantiene exactamente como en el script original:
    - mat_select: "Semana 14"
    - checkbox: "14-input"

    Esto conserva las mismas acciones de descarga, botones y selección interna
    que ya funcionaban en RUN_RENOMBRAR_MIN.py.
    """
    global FECHA_DESDE, FECHA_HASTA, SEMANA_SELECCIONADA, ARCHIVO_SALIDA

    SEMANA_SELECCIONADA = int(numero_semana)
    inicio, fin = calcular_rango_semana(numero_semana)
    FECHA_DESDE = inicio.strftime("%d-%m-%Y")
    FECHA_HASTA = fin.strftime("%d-%m-%Y")
    try:
        ARCHIVO_SALIDA = CARPETA_SALIDA / f"Semana {SEMANA_SELECCIONADA}.xlsx"
    except Exception:
        pass

    return inicio, fin



def aplicar_dia_a_config(dia: str):
    """Actualiza únicamente el comportamiento del consolidado: lunes crea, martes a domingo anexa."""
    global DIA_SELECCIONADO
    DIA_SELECCIONADO = str(dia).strip()
    return DIA_SELECCIONADO


def configurar_rutas(carpeta_base: str):
    """Ajusta rutas para descarga y consolidado. Por default usa Desktop/Renovaciones y Desktop/Renovaciones/Renovacion."""
    global DOWNLOAD_DIR, CONSOLIDADO_DIR, ANTERIORES_DIR, PREFS
    global CARPETA_ENTRADA, CARPETA_SALIDA, ARCHIVO_SALIDA

    base = Path(carpeta_base).expanduser().resolve()
    salida = base / "Renovacion"

    base.mkdir(parents=True, exist_ok=True)
    salida.mkdir(parents=True, exist_ok=True)

    DOWNLOAD_DIR = str(base)
    CONSOLIDADO_DIR = str(base)
    ANTERIORES_DIR = str(base)

    PREFS["download.default_directory"] = DOWNLOAD_DIR

    CARPETA_ENTRADA = base
    CARPETA_SALIDA = salida
    ARCHIVO_SALIDA = salida / f"Semana {SEMANA_SELECCIONADA}.xlsx"

    return base, salida


def ejecutar_descarga_todo():
    """Replica la opción 0 del script original: primero bloque prioritario y luego marcas finales."""
    print("\n>>> LIMPIANDO CARPETA DE DESCARGAS...")
    limpiar_carpeta_descargas()

    print("\n>>> INICIANDO BLOQUE PRIORITARIO...")
    for s in SITES:
        revisar_pausa()
        marcas_filtradas = [
            m for m in s["logos"]
            if not (
                (s["name"] == "GUATEMALA" or s["name"] == "PERU")
                and m == "PRESICO"
            )
        ]
        if marcas_filtradas:
            ejecutar_sitio(s, marcas_filtradas)

    print("\n>>> EJECUTANDO MARCAS FINALES...")
    for pais in ["GUATEMALA", "PERU"]:
        revisar_pausa()
        cfg = next((s for s in SITES if s["name"] == pais), None)
        if cfg:
            ejecutar_sitio(cfg, ["PRESICO"])


def ejecutar_proceso_completo():
    ejecutar_descarga_todo()
    print("\n>>> INICIANDO CONSOLIDADO...")
    ejecutar_consolidado()


def comprimir_csvs(carpeta: Path):
    csvs = sorted(carpeta.glob("*.csv"))
    if not csvs:
        return None

    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for archivo in csvs:
            zf.write(archivo, arcname=archivo.name)
    buffer.seek(0)
    return buffer


def mostrar_descargas(carpeta_base: Path):
    archivo_excel = carpeta_base / "Renovacion" / f"Semana {SEMANA_SELECCIONADA}.xlsx"
    if archivo_excel.exists():
        st.download_button(
            f"Descargar consolidado Semana {SEMANA_SELECCIONADA}.xlsx",
            data=archivo_excel.read_bytes(),
            file_name=f"Semana {SEMANA_SELECCIONADA}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    zip_csvs = comprimir_csvs(carpeta_base)
    if zip_csvs is not None:
        st.download_button(
            "Descargar CSV descargados en ZIP",
            data=zip_csvs,
            file_name="Renovaciones_CSV.zip",
            mime="application/zip",
        )


def main_app_renovaciones():
    st.title("Renovaciones LATAM")
    st.title("Renovaciones LATAM")

    st.write("Selecciona qué quieres ejecutar. Siempre se solicita la semana y se calcula el rango tomando como base que la semana 24 es del 08-06-2026 al 14-06-2026.")

    accion = st.selectbox(
        "¿Qué quieres hacer?",
        [
            "Ejecutar proceso completo",
            "Solo descargar archivos",
            "Solo hacer consolidado",
        ],
    )

    semana = st.number_input("Semana", min_value=1, max_value=60, value=24, step=1)
    inicio, fin = calcular_rango_semana(int(semana))
    st.info(f"Semana {int(semana)}: {inicio.strftime('%d-%m-%Y')} al {fin.strftime('%d-%m-%Y')}")

    dia = st.selectbox(
        "¿Qué día es?",
        ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"],
        index=0,
    )
    if dia == "Lunes":
        st.caption("Lunes: se genera el consolidado semanal desde cero.")
    else:
        st.caption(f"{dia}: se abre el archivo de la semana y se pega abajo la nueva información.")

    carpeta_default = RUTA_BASE_DEFAULT
    carpeta_base_txt = st.text_input("Carpeta de trabajo", value=carpeta_default)
    st.caption(f"Los CSV se guardan en: {carpeta_base_txt} | El consolidado se guarda en: {Path(carpeta_base_txt) / 'Renovacion'}")

    if "pausar_proceso_renovaciones" not in st.session_state:
        st.session_state["pausar_proceso_renovaciones"] = False

    st.subheader("Control del proceso")
    col_pausa, col_reanuda = st.columns(2)
    with col_pausa:
        if st.button("Pausar proceso"):
            st.session_state["pausar_proceso_renovaciones"] = True
    with col_reanuda:
        if st.button("Reanudar proceso"):
            st.session_state["pausar_proceso_renovaciones"] = False

    if st.session_state.get("pausar_proceso_renovaciones", False):
        st.warning("⏸️ Pausa activada. El proceso se detendrá entre pasos y continuará al presionar Reanudar proceso.")
    else:
        st.success("▶️ Listo para ejecutar / proceso sin pausa.")

    with st.expander("Configuración Selenium", expanded=False):
        global CHROME_BIN, DRIVER_PATH, USERNAME, PASSWORD
        CHROME_BIN = st.text_input("Ruta de Chrome", value=CHROME_BIN)
        driver_txt = st.text_input("Ruta de ChromeDriver opcional", value=DRIVER_PATH or "")
        DRIVER_PATH = driver_txt.strip() or None
        USERNAME = st.text_input("Usuario", value=USERNAME)
        PASSWORD = st.text_input("Contraseña", value=PASSWORD, type="password")

    if st.button("Ejecutar", type="primary"):
        st.session_state["pausar_proceso_renovaciones"] = False
        aplicar_semana_a_config(int(semana))
        aplicar_dia_a_config(dia)
        carpeta_base, _ = configurar_rutas(carpeta_base_txt)
        inicializar_estado_descargas(accion)

        buffer = io.StringIO()
        ok = True

        with st.spinner("Ejecutando proceso..."):
            try:
                with contextlib.redirect_stdout(buffer), contextlib.redirect_stderr(buffer):
                    print(f"Acción seleccionada: {accion}")
                    print(f"Semana seleccionada: {int(semana)}")
                    print(f"Día seleccionado: {DIA_SELECCIONADO}")
                    print(f"Rango usado: {FECHA_DESDE} al {FECHA_HASTA}")
                    print(f"Carpeta usada: {carpeta_base}")
                    print(f"Consolidado de salida: {ARCHIVO_SALIDA}")

                    if accion == "Ejecutar proceso completo":
                        ejecutar_proceso_completo()
                    elif accion == "Solo descargar archivos":
                        ejecutar_descarga_todo()
                    elif accion == "Solo hacer consolidado":
                        ejecutar_consolidado()

            except Exception as e:
                ok = False
                print(f"\n❌ Error general: {e}")
                traceback.print_exc(file=buffer)

        if ok:
            st.success("Proceso terminado.")
        else:
            st.error("El proceso terminó con error. Revisa el log.")

        st.subheader("Log")
        st.text_area("Salida del proceso", value=buffer.getvalue(), height=420)

        mostrar_descargas(carpeta_base)

    else:
        carpeta_base, _ = configurar_rutas(carpeta_base_txt)
        mostrar_descargas(carpeta_base)


# ============================================================
# TABLERO CARTERAS: CONVERTIR EXCEL PRESICO A PARQUET
# ============================================================

RUTA_DESCARGAS_CARTERAS_DEFAULT = r"C:\Users\EQUIPO\Downloads"
NOMBRE_EXCEL_PRESICO_DEFAULT = "PRESICO 06-06-2026.xlsx"


def extraer_fecha_nombre_presico(nombre_archivo: str):
    """Extrae la fecha desde nombres tipo PRESICO 07-06-2026.xlsx."""
    m = re.search(r"PRESICO\s+(\d{2})-(\d{2})-(\d{4})\.xlsx$", str(nombre_archivo), flags=re.IGNORECASE)
    if not m:
        return None
    dia, mes, anio = map(int, m.groups())
    return _dt.date(anio, mes, dia)


def nombre_presico_por_fecha(fecha: _dt.date, extension: str):
    return f"PRESICO {fecha.strftime('%d-%m-%Y')}{extension}"


def preparar_df_cartera_para_parquet(ruta_excel: Path, fecha_corte_forzada=None, status_callback=None):
    """
    Replica la lógica original de conversión a parquet:
    - lee el Excel
    - elimina ruta con ZONA DE PRUEBAS-
    - convierte Localidad y objetos a texto
    - opcionalmente cambia la columna Corte a una fecha específica
    """
    if status_callback:
        status_callback(f"Leyendo Excel: {ruta_excel.name}")

    print(f"Leyendo Excel: {ruta_excel}")
    df = pd.read_excel(ruta_excel)

    total_inicial = len(df)
    columnas_iniciales = len(df.columns)
    conteo_eliminadas = 0

    print(f"Filas totales cargadas: {total_inicial}")
    print(f"Columnas totales cargadas: {columnas_iniciales}")

    if status_callback:
        status_callback("Revisando y eliminando filas de ZONA DE PRUEBAS-")

    if 'ruta' in df.columns:
        filas_a_eliminar = df['ruta'].astype(str).str.contains('ZONA DE PRUEBAS-', na=False)
        conteo_eliminadas = int(filas_a_eliminar.sum())

        if conteo_eliminadas > 0:
            df = df[~filas_a_eliminar].copy()
            print(f"🔍 Se eliminaron {conteo_eliminadas} filas de 'ZONA DE PRUEBAS-'.")
        else:
            print("✅ No se encontraron filas de prueba.")
    else:
        print("⚠️ Advertencia: Columna 'ruta' no encontrada.")

    if fecha_corte_forzada is not None:
        if status_callback:
            status_callback(f"Actualizando columna Corte a {fecha_corte_forzada.strftime('%d/%m/%Y')}")

        if 'Corte' in df.columns:
            df['Corte'] = pd.Timestamp(fecha_corte_forzada)
            print(f"✅ Columna Corte actualizada a: {fecha_corte_forzada.strftime('%d/%m/%Y')}")
        else:
            print("⚠️ Advertencia: Columna 'Corte' no encontrada. No se cambió fecha de corte.")

    if status_callback:
        status_callback("Ajustando tipos de datos antes de guardar el parquet")

    if 'Localidad' in df.columns:
        df['Localidad'] = df['Localidad'].astype(str)

    for col in df.select_dtypes(include=['object']).columns:
        df[col] = df[col].astype(str)

    columnas_finales = len(df.columns)

    print(f"Filas finales: {len(df)}")
    print(f"Columnas finales: {columnas_finales}")

    info = {
        "filas_iniciales": int(total_inicial),
        "filas_finales": int(len(df)),
        "filas_eliminadas_prueba": int(conteo_eliminadas),
        "columnas_iniciales": int(columnas_iniciales),
        "columnas_finales": int(columnas_finales),
        "mantiene_eliminacion_zona_pruebas": True,
    }

    return df, info


def guardar_parquet_cartera(df: pd.DataFrame, ruta_salida: Path, status_callback=None):
    ruta_salida.parent.mkdir(parents=True, exist_ok=True)

    if status_callback:
        status_callback(f"Guardando parquet: {ruta_salida.name}")

    print(f"🚀 Guardando archivo Parquet: {ruta_salida}")
    df.to_parquet(ruta_salida, engine='pyarrow', index=False)
    print(f"✅ Proceso completado. Guardado en: {ruta_salida}")


def _resultado_parquet(nombre: str, ruta_salida: Path, df: pd.DataFrame, info: dict):
    resultado = dict(info)
    resultado.update({
        "nombre": nombre,
        "ruta": str(ruta_salida),
        "archivo": ruta_salida.name,
        "preview": df.iloc[:5, :2].copy(),
    })
    return resultado


def convertir_excel_a_parquet_carteras(ruta_excel_presico: str, dia: str, status_callback=None):
    """
    Lunes:
    - crea un parquet con el archivo original y el nombre del domingo.
    - crea otro parquet con Corte cambiado al sábado anterior y nombre del sábado.

    Martes a domingo:
    - crea solo un parquet con el mismo nombre/fecha del Excel.
    """
    ruta_excel = Path(ruta_excel_presico)
    print("\n--- [INICIO] TAREA: CONVERSIÓN CARTERA A PARQUET ---")

    if status_callback:
        status_callback("Validando que exista el archivo Excel seleccionado")

    if not ruta_excel.exists():
        print(f"❌ Error: No existe el archivo {ruta_excel}")
        return []

    fecha_archivo = extraer_fecha_nombre_presico(ruta_excel.name)
    if fecha_archivo is None:
        print("⚠️ No se pudo extraer fecha del nombre. Se generará parquet con el mismo nombre del Excel.")
        df, info = preparar_df_cartera_para_parquet(ruta_excel, status_callback=status_callback)
        salida = ruta_excel.with_suffix('.parquet')
        guardar_parquet_cartera(df, salida, status_callback=status_callback)
        return [_resultado_parquet("Parquet generado", salida, df, info)]

    dia_normalizado = str(dia).strip().lower()
    resultados = []

    if dia_normalizado == "lunes":
        print("Día seleccionado: Lunes")
        print("Se generarán dos parquet: domingo sin modificar y sábado con Corte ajustado.")

        if status_callback:
            status_callback("Generando parquet del domingo sin modificar")

        # 1) Parquet del domingo sin cambiar Corte.
        df_domingo, info_domingo = preparar_df_cartera_para_parquet(ruta_excel, status_callback=status_callback)
        salida_domingo = ruta_excel.parent / nombre_presico_por_fecha(fecha_archivo, ".parquet")
        guardar_parquet_cartera(df_domingo, salida_domingo, status_callback=status_callback)
        resultados.append(_resultado_parquet("Domingo sin modificar", salida_domingo, df_domingo, info_domingo))

        if status_callback:
            status_callback("Generando parquet del sábado anterior con Corte ajustado")

        # 2) Parquet del sábado anterior, cambiando Corte.
        fecha_sabado = fecha_archivo - _dt.timedelta(days=1)
        df_sabado, info_sabado = preparar_df_cartera_para_parquet(ruta_excel, fecha_corte_forzada=fecha_sabado, status_callback=status_callback)
        salida_sabado = ruta_excel.parent / nombre_presico_por_fecha(fecha_sabado, ".parquet")
        guardar_parquet_cartera(df_sabado, salida_sabado, status_callback=status_callback)
        resultados.append(_resultado_parquet("Sábado con Corte ajustado", salida_sabado, df_sabado, info_sabado))

    else:
        print(f"Día seleccionado: {dia}")
        print("Se generará un parquet con el mismo nombre/fecha del Excel.")

        if status_callback:
            status_callback(f"Generando parquet de {dia} con el mismo nombre/fecha del Excel")

        df, info = preparar_df_cartera_para_parquet(ruta_excel, status_callback=status_callback)
        salida = ruta_excel.parent / nombre_presico_por_fecha(fecha_archivo, ".parquet")
        guardar_parquet_cartera(df, salida, status_callback=status_callback)
        resultados.append(_resultado_parquet("Parquet generado", salida, df, info))

    if status_callback:
        status_callback("Proceso de parquet terminado")

    print("\n🏁 Tarea de parquet finalizada.")
    return resultados


def _leer_preview_parquet_seguro(ruta: Path, filas: int = 5, columnas: int = 2):
    """Lee solo las primeras columnas y primeras filas del parquet para que la vista previa sea ligera."""
    try:
        if not ruta.exists():
            return None, f"No encontré el archivo generado en: {ruta}"

        # Primero leemos únicamente la metadata para tomar solo las primeras columnas.
        # Esto evita cargar las 46 columnas completas solo para la vista previa.
        try:
            import pyarrow.parquet as pq
            pf = pq.ParquetFile(str(ruta))
            columnas_disponibles = pf.schema.names
            columnas_preview = columnas_disponibles[:columnas]
            if not columnas_preview:
                return None, "El parquet no tiene columnas disponibles para vista previa."
            df_preview = pd.read_parquet(ruta, columns=columnas_preview).head(filas)
        except Exception:
            # Respaldo si pyarrow no permite leer metadata: carga normal, pero recorta de inmediato.
            df_preview = pd.read_parquet(ruta).iloc[:filas, :columnas]

        # Para evitar problemas de renderizado por tipos mixtos, fechas o nulos,
        # solo la vista previa se convierte a texto. El parquet queda intacto.
        df_preview = df_preview.copy()
        for col in df_preview.columns:
            df_preview[col] = df_preview[col].apply(lambda x: "" if pd.isna(x) else str(x))

        return df_preview, None
    except Exception as e:
        return None, str(e)


def mostrar_resultados_parquet(resultados):
    if not resultados:
        return

    st.subheader("Vista previa de archivos generados")
    st.info("Se mantiene la eliminación de filas donde la columna ruta contiene 'ZONA DE PRUEBAS-'.")

    for idx, resultado in enumerate(resultados):
        ruta = Path(resultado.get("ruta", ""))

        with st.container(border=True):
            st.markdown(f"**{resultado.get('nombre', 'Parquet')}**")
            st.caption(resultado.get("archivo", ruta.name))

            m1, m2, m3 = st.columns(3)
            m1.metric("Filas iniciales", f"{resultado.get('filas_iniciales', 0):,}")
            m2.metric("Filas eliminadas", f"{resultado.get('filas_eliminadas_prueba', 0):,}")
            m3.metric("Filas finales", f"{resultado.get('filas_finales', 0):,}")

            c1, c2 = st.columns(2)
            c1.metric("Columnas iniciales", f"{resultado.get('columnas_iniciales', 0):,}")
            c2.metric("Columnas finales", f"{resultado.get('columnas_finales', 0):,}")

            st.markdown("**Vista previa ligera**")

            preview, error_preview = _leer_preview_parquet_seguro(ruta, filas=5, columnas=2)

            if preview is not None and not preview.empty:
                st.caption("Mostrando solo las primeras 2 columnas y las primeras 5 filas para que cargue rápido.")
                st.table(preview)
            elif error_preview:
                st.warning(f"No se pudo cargar la vista previa de {ruta.name}: {error_preview}")
            else:
                st.warning("La vista previa no tiene filas para mostrar, pero el archivo se generó correctamente.")

            if ruta.exists():
                st.download_button(
                    f"Descargar {ruta.name}",
                    data=ruta.read_bytes(),
                    file_name=ruta.name,
                    mime="application/octet-stream",
                    key=f"download_parquet_{idx}_{ruta.name}",
                )
            else:
                st.warning(f"No encontré el archivo para descargar: {ruta}")

def archivos_excel_presico_en_descargas(carpeta_descargas: Path):
    if not carpeta_descargas.exists():
        return []
    return sorted(carpeta_descargas.glob("PRESICO *.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)


def main_app_carteras():
    st.title("Carteras")
    st.write("Convierte el Excel de PRESICO a Parquet. El flujo LATAM/OneDrive no se agregó.")

    carpeta_descargas_txt = st.text_input("Carpeta de descargas", value=RUTA_DESCARGAS_CARTERAS_DEFAULT)
    carpeta_descargas = Path(carpeta_descargas_txt).expanduser().resolve()

    archivos = archivos_excel_presico_en_descargas(carpeta_descargas)
    nombres_archivos = [p.name for p in archivos]

    default_path = carpeta_descargas / NOMBRE_EXCEL_PRESICO_DEFAULT

    if nombres_archivos:
        opciones = nombres_archivos.copy()
        if NOMBRE_EXCEL_PRESICO_DEFAULT not in opciones:
            opciones.insert(0, NOMBRE_EXCEL_PRESICO_DEFAULT)
        index_default = opciones.index(NOMBRE_EXCEL_PRESICO_DEFAULT) if NOMBRE_EXCEL_PRESICO_DEFAULT in opciones else 0
        nombre_seleccionado = st.selectbox("Archivo PRESICO a buscar en descargas", opciones, index=index_default)
        ruta_excel_txt = str(carpeta_descargas / nombre_seleccionado)
    else:
        st.warning("No encontré archivos 'PRESICO *.xlsx' en la carpeta. Puedes dejar o escribir el nombre esperado.")
        nombre_seleccionado = st.text_input("Nombre del archivo PRESICO", value=NOMBRE_EXCEL_PRESICO_DEFAULT)
        ruta_excel_txt = str(carpeta_descargas / nombre_seleccionado)

    ruta_excel_txt = st.text_input("Ruta completa del Excel", value=ruta_excel_txt)

    dia = st.selectbox(
        "¿Qué día es?",
        ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"],
        index=0,
        key="dia_carteras",
    )

    if dia == "Lunes":
        st.caption("Lunes: genera el parquet del domingo sin modificar y otro del sábado anterior cambiando la columna Corte.")
    else:
        st.caption(f"{dia}: genera un parquet con el mismo nombre/fecha del Excel seleccionado.")

    if "carteras_resultados_parquet" not in st.session_state:
        st.session_state["carteras_resultados_parquet"] = []
    if "carteras_log_parquet" not in st.session_state:
        st.session_state["carteras_log_parquet"] = ""

    if st.button("Ejecutar Carteras", type="primary"):
        buffer = io.StringIO()
        resultados = []
        ok = True
        estado_box = st.empty()

        def actualizar_estado(mensaje):
            estado_box.info(f"🔄 {mensaje}")

        with st.spinner("Procesando cartera..."):
            try:
                with contextlib.redirect_stdout(buffer), contextlib.redirect_stderr(buffer):
                    resultados = convertir_excel_a_parquet_carteras(
                        ruta_excel_txt,
                        dia,
                        status_callback=actualizar_estado,
                    )
            except Exception as e:
                ok = False
                print(f"\n❌ Error general en Carteras: {e}")
                traceback.print_exc(file=buffer)

        st.session_state["carteras_resultados_parquet"] = resultados
        st.session_state["carteras_log_parquet"] = buffer.getvalue()

        if ok:
            estado_box.success("✅ Proceso de parquet terminado")
            st.success("Proceso de Carteras terminado.")
        else:
            estado_box.error("❌ El proceso de parquet terminó con error")
            st.error("El proceso de Carteras terminó con error. Revisa el log.")

    # El log técnico de Carteras se conserva internamente, pero ya no se muestra en pantalla
    # para dejar limpia la vista de resultados.
    mostrar_resultados_parquet(st.session_state.get("carteras_resultados_parquet", []))


def main_app():
    st.set_page_config(page_title="Gestor de Reportes PRESICO", layout="wide")

    st.sidebar.title("Gestor de Reportes PRESICO")
    tablero = st.sidebar.radio(
        "¿Qué tablero quieres abrir?",
        ["Carteras", "Renovaciones"],
        index=0,
    )

    if tablero == "Carteras":
        main_app_carteras()
    else:
        main_app_renovaciones()


if __name__ == "__main__":
    main_app()
